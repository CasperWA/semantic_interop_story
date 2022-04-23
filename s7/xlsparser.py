from pydantic.dataclasses import dataclass
from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange
from itertools import product
import numpy as np
import re

from .models import Workbook


@dataclass
class XLSParser:
    """Parse excel documents given document configuration
    and an S7-entity
    """

    xls_model: Workbook

    def get(self):
        wb_obj = load_workbook(self.xls_model.workbook.path, data_only=True)
        sheet_obj = wb_obj[self.xls_model.sheet]

        alldata = {}
        redata = {}
        for d in self.xls_model.data:
            item = []
            dataRange = CellRange(self.xls_model.data_range)
            cellRange = dataRange.intersection(CellRange(d[1].range))
            for cell in product(
                range(cellRange.min_row, cellRange.max_row + 1),
                range(cellRange.min_col, cellRange.max_col + 1),
            ):
                cell_obj = sheet_obj.cell(*cell)
                v = str(cell_obj.value)
                if d[1].type == "float":
                    try:
                        item.append(float(v))
                        # print ("added ", float(v))
                    except ValueError:
                        item.append(v)
                else:
                    item.append(v)

            alldata[d[0]] = item
            redata[d[0]] = d[1].regexp

        newdata = {}
        for elem in np.column_stack([alldata[label] for label in alldata.keys()]):
            l = list(zip([redata[label] for label in alldata.keys()], elem))
            if not False in [bool(re.match(*s)) for s in l]:
                for (k, v) in list(zip(alldata.keys(), elem)):
                    if not k in newdata:
                        newdata[k] = []
                    newdata[k].append(v)
        return newdata
